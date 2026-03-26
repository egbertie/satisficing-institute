#!/usr/bin/env python3
"""
统一文件处理器 - Universal File Handler
整合所有文件处理能力：DOCX、PDF、Excel、CSV等
"""

import os
import sys
import subprocess
import tempfile
from pathlib import Path
from typing import Optional

# 支持的文件类型
SUPPORTED_TYPES = {
    '.md': 'markdown',
    '.txt': 'text',
    '.docx': 'docx',
    '.pdf': 'pdf',
    '.xlsx': 'excel',
    '.csv': 'csv',
    '.png': 'image',
    '.jpg': 'image',
    '.jpeg': 'image',
    '.gif': 'image',
}

class FileHandler:
    """统一文件处理器"""
    
    def __init__(self):
        self.workspace = Path("/root/.openclaw/workspace")
        
    def read_file(self, file_path: str) -> str:
        """
        读取任何支持的文件类型并返回文本内容
        
        Args:
            file_path: 文件路径
            
        Returns:
            文件文本内容
        """
        path = Path(file_path)
        if not path.exists():
            return f"Error: File not found: {file_path}"
        
        ext = path.suffix.lower()
        
        if ext == '.md' or ext == '.txt':
            return self._read_text(path)
        elif ext == '.docx':
            return self._read_docx(path)
        elif ext == '.pdf':
            return self._read_pdf(path)
        elif ext == '.xlsx':
            return self._read_excel(path)
        elif ext == '.csv':
            return self._read_csv(path)
        else:
            return f"Error: Unsupported file type: {ext}"
    
    def _read_text(self, path: Path) -> str:
        """读取文本文件"""
        with open(path, 'r', encoding='utf-8') as f:
            return f.read()
    
    def _read_docx(self, path: Path) -> str:
        """读取DOCX文件 - 使用pandoc转换为markdown"""
        try:
            # 使用pandoc转换
            result = subprocess.run(
                ['pandoc', str(path), '-t', 'markdown'],
                capture_output=True,
                text=True,
                timeout=30
            )
            if result.returncode == 0:
                return result.stdout
            else:
                return f"Error converting DOCX: {result.stderr}"
        except Exception as e:
            return f"Error reading DOCX: {str(e)}"
    
    def _read_pdf(self, path: Path) -> str:
        """读取PDF文件 - 使用pypdf"""
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(path))
            text_parts = []
            for i, page in enumerate(reader.pages):
                text = page.extract_text()
                if text:
                    text_parts.append(f"\n--- Page {i+1} ---\n{text}")
            return "\n".join(text_parts)
        except ImportError:
            return "Error: pypdf not installed. Run: pip install pypdf"
        except Exception as e:
            return f"Error reading PDF: {str(e)}"
    
    def _read_excel(self, path: Path) -> str:
        """读取Excel文件 - 转换为CSV格式文本"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path))
            result = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                result.append(f"\n=== Sheet: {sheet_name} ===")
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell else "" for cell in row])
                    result.append(row_text)
            return "\n".join(result)
        except ImportError:
            return "Error: openpyxl not installed. Run: pip install openpyxl"
        except Exception as e:
            return f"Error reading Excel: {str(e)}"
    
    def _read_csv(self, path: Path) -> str:
        """读取CSV文件"""
        try:
            import csv
            with open(path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows = list(reader)
                # 格式化输出
                result = []
                for i, row in enumerate(rows[:100]):  # 限制100行
                    result.append(" | ".join(row))
                if len(rows) > 100:
                    result.append(f"\n... ({len(rows) - 100} more rows)")
                return "\n".join(result)
        except Exception as e:
            return f"Error reading CSV: {str(e)}"
    
    def convert_file(self, input_path: str, output_format: str) -> str:
        """
        转换文件格式
        
        Args:
            input_path: 输入文件路径
            output_format: 输出格式 (md, txt, docx, pdf, etc.)
            
        Returns:
            输出文件路径
        """
        path = Path(input_path)
        if not path.exists():
            return f"Error: File not found: {input_path}"
        
        # 创建临时输出路径
        output_path = Path(tempfile.gettempdir()) / f"{path.stem}.{output_format}"
        
        input_ext = path.suffix.lower()
        
        # DOCX -> Markdown
        if input_ext == '.docx' and output_format == 'md':
            return self._convert_docx_to_md(path, output_path)
        
        # Markdown -> DOCX
        elif input_ext in ['.md', '.txt'] and output_format == 'docx':
            return self._convert_md_to_docx(path, output_path)
        
        # Excel -> CSV
        elif input_ext == '.xlsx' and output_format == 'csv':
            return self._convert_excel_to_csv(path, output_path)
        
        # CSV -> Excel
        elif input_ext == '.csv' and output_format == 'xlsx':
            return self._convert_csv_to_excel(path, output_path)
        
        else:
            return f"Error: Conversion from {input_ext} to {output_format} not supported"
    
    def _convert_docx_to_md(self, input_path: Path, output_path: Path) -> str:
        """DOCX转Markdown"""
        try:
            subprocess.run(
                ['pandoc', str(input_path), '-o', str(output_path)],
                check=True,
                timeout=30
            )
            return str(output_path)
        except Exception as e:
            return f"Error: {str(e)}"
    
    def _convert_md_to_docx(self, input_path: Path, output_path: Path) -> str:
        """Markdown转DOCX"""
        try:
            subprocess.run(
                ['pandoc', str(input_path), '-o', str(output_path)],
                check=True,
                timeout=30
            )
            return str(output_path)
        except Exception as e:
            return f"Error: {str(e)}"
    
    def _convert_excel_to_csv(self, input_path: Path, output_path: Path) -> str:
        """Excel转CSV"""
        try:
            import openpyxl
            import csv
            
            wb = openpyxl.load_workbook(str(input_path))
            sheet = wb.active
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow(row)
            
            return str(output_path)
        except Exception as e:
            return f"Error: {str(e)}"
    
    def _convert_csv_to_excel(self, input_path: Path, output_path: Path) -> str:
        """CSV转Excel"""
        try:
            import openpyxl
            import csv
            
            wb = openpyxl.Workbook()
            sheet = wb.active
            
            with open(input_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    sheet.append(row)
            
            wb.save(str(output_path))
            return str(output_path)
        except Exception as e:
            return f"Error: {str(e)}"
    
    def upload_file(self, file_path: str, channel: str = "feishu") -> str:
        """
        上传文件到指定渠道
        
        Args:
            file_path: 文件路径
            channel: 渠道名称 (feishu, notion, email, telegram)
            
        Returns:
            上传结果
        """
        gateway_script = self.workspace / "skills" / "file-gateway" / "scripts" / "file_gateway.py"
        
        if not gateway_script.exists():
            return "Error: file-gateway not installed"
        
        try:
            result = subprocess.run(
                ['python3', str(gateway_script), 'upload', file_path, '--channel', channel],
                capture_output=True,
                text=True,
                timeout=60
            )
            return result.stdout if result.returncode == 0 else result.stderr
        except Exception as e:
            return f"Error uploading file: {str(e)}"


def main():
    """命令行入口"""
    if len(sys.argv) < 3:
        print("Usage: file_handler.py [read|convert|upload] <file> [options]")
        print("\nExamples:")
        print("  file_handler.py read document.docx")
        print("  file_handler.py convert document.docx md")
        print("  file_handler.py upload document.pdf feishu")
        sys.exit(1)
    
    handler = FileHandler()
    command = sys.argv[1]
    file_path = sys.argv[2]
    
    if command == "read":
        content = handler.read_file(file_path)
        print(content)
    
    elif command == "convert":
        if len(sys.argv) < 4:
            print("Error: Please specify output format")
            sys.exit(1)
        output_format = sys.argv[3]
        result = handler.convert_file(file_path, output_format)
        print(result)
    
    elif command == "upload":
        channel = sys.argv[3] if len(sys.argv) > 3 else "feishu"
        result = handler.upload_file(file_path, channel)
        print(result)
    
    else:
        print(f"Unknown command: {command}")
        sys.exit(1)


if __name__ == "__main__":
    main()
